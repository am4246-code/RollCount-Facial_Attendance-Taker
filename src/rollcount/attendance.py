from __future__ import annotations

import logging
import tkinter as tk
import calendar
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from datetime import timedelta, time
from tkinter import messagebox, ttk


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _auto_size_columns(worksheet): # Revert: Move back inside export_daily_summary
    """Adjusts column widths to fit content."""
    for col in worksheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError: # Handle non-string types
                pass
        adjusted_width = (max_length + 2) * 1.2 # Add a little padding
        worksheet.column_dimensions[column].width = adjusted_width
# Revert: Move Student import back to top
from .registry import Student
from .config import ATTENDANCE_LOG_DIR, DAILY_LOG_DIR, WEEKLY_LOG_DIR, MONTHLY_LOG_DIR, DEFAULT_ATTENDANCE_FILE, ensure_directories


logger = logging.getLogger(__name__)


@dataclass
class AttendanceRecord:
    """Record of a student's attendance at a specific time."""
    student_id: str
    full_name: str
    recorded_at: datetime
    status: str = "Present"


class AttendanceLogger:
    """Manages attendance logging to Excel workbooks."""
    
    def __init__(self, workbook_path: Path = DEFAULT_ATTENDANCE_FILE) -> None:
        """Initialize the attendance logger.
        
        Args:
            workbook_path: Path to the attendance Excel workbook
        """
        self.workbook_path = workbook_path
        ensure_directories()
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        """Ensure the attendance workbook exists and is properly formatted."""
        if self.workbook_path.exists():
            return
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            if worksheet is None:
                worksheet = workbook.create_sheet()
            worksheet.title = "Attendance"
            worksheet.append(["Date", "Time", "Student ID", "Full Name", "Status"])
            workbook.save(self.workbook_path)
            logger.info(f"Created new attendance workbook at {self.workbook_path}")
        except OSError as e:
            logger.error(f"Failed to create attendance workbook: {e}")
            raise

    def records_for_day(self, day: datetime) -> list[AttendanceRecord]:
        """Get all attendance records for a specific day.
        
        Args:
            day: Date to retrieve records for
            
        Returns:
            List of AttendanceRecord objects for the day, sorted by time
        """
        try:
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active
            if worksheet is None:
                return []
                
            day_text = day.strftime("%Y-%m-%d")
            records: list[AttendanceRecord] = []
            
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row is None or len(row) < 5:
                    continue
                row_date, row_time, row_student_id, row_full_name, row_status = row
                # Convert row_date to string if it's a datetime object
                if isinstance(row_date, datetime):
                    row_date_str = row_date.strftime("%Y-%m-%d")
                else:
                    row_date_str = str(row_date) if row_date else ""
                if row_date_str != day_text:
                    continue
                try:
                    recorded_at = datetime.strptime(f"{row_date_str} {row_time}", "%Y-%m-%d %H:%M:%S")
                    records.append(
                        AttendanceRecord(
                            student_id=str(row_student_id),
                            full_name=str(row_full_name),
                            recorded_at=recorded_at,
                            status=str(row_status or "Present"),
                        )
                    )
                except (ValueError, TypeError) as e:
                    logger.warning(f"Skipping malformed attendance record: {e}")
                    continue
                    
            return sorted(records, key=lambda record: record.recorded_at)
        except OSError as e:
            logger.error(f"Error retrieving records for day: {e}")
            return []

    def _get_presence_for_range(self, start_date: datetime, end_date: datetime) -> dict[str, set[str]]:
        """
        Loads attendance for a given date range into an efficient lookup structure.

        Args:
            start_date: The start date (inclusive).
            end_date: The end date (inclusive).

        Returns:
            A dictionary where keys are date strings ('YYYY-MM-DD') and
            values are sets of student IDs present on that day.
        """
        presence_by_day: dict[str, set[str]] = {}
        try:
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active
            if worksheet is None:
                return {}

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3:
                    continue
                row_date_str, _, student_id, *_ = row
                try:
                    if isinstance(row_date_str, datetime):
                        row_date = row_date_str
                    else:
                        row_date = datetime.strptime(str(row_date_str), "%Y-%m-%d")
                    if start_date.date() <= row_date.date() <= end_date.date() and student_id is not None:
                        date_key = row_date.strftime("%Y-%m-%d")
                        if date_key not in presence_by_day:
                            presence_by_day[date_key] = set()
                        presence_by_day[date_key].add(str(student_id))
                except (ValueError, TypeError):
                    continue
        except OSError:
            logger.error(f"Could not read workbook to build presence for range {start_date.date()} to {end_date.date()}")
        return presence_by_day

    def _get_monthly_presence(self, year: int, month: int) -> dict[str, set[str]]:
        """
        Loads all attendance for a given month into an efficient lookup structure.

        Args:
            year: The year of the month to load.
            month: The month number (1-12).

        Returns:
            A dictionary where keys are date strings ('YYYY-MM-DD') and
            values are sets of student IDs present on that day.
        """
        presence_by_day: dict[str, set[str]] = {}
        try:
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active
            if worksheet is None:
                return {}

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3:
                    continue
                row_date_str, _, student_id, *_ = row
                try:
                    if isinstance(row_date_str, datetime):
                        row_date = row_date_str
                    else:
                        row_date = datetime.strptime(str(row_date_str), "%Y-%m-%d")
                    if row_date.year == year and row_date.month == month and student_id is not None:
                        date_key = row_date.strftime("%Y-%m-%d")
                        if date_key not in presence_by_day:
                            presence_by_day[date_key] = set()
                        presence_by_day[date_key].add(str(student_id))
                except (ValueError, TypeError):
                    continue
        except OSError:
            logger.error(f"Could not read workbook to build monthly presence for {year}-{month}")
        return presence_by_day

    def append_record(self, record: AttendanceRecord) -> bool:
        """Add a new attendance record to the workbook.
        
        Args:
            record: AttendanceRecord to add
            
        """
        try:
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active
            if worksheet is None:
                logger.error("Workbook has no active sheet")
                return False
                
            worksheet.append(
                [
                    record.recorded_at.strftime("%Y-%m-%d"),
                    record.recorded_at.strftime("%H:%M:%S"),
                    record.student_id,
                    record.full_name,
                    record.status,
                ]
            )
            workbook.save(self.workbook_path)
            logger.debug(f"Added attendance record for {record.student_id}")
            return True
        except OSError as e:
            logger.error(f"Error appending record: {e}")
            return False

    def clear_records_for_day(self, day: datetime) -> None:
        """Remove all attendance records for a specific day.
        
        Args:
            day: Date to clear records for
        """
        try:
            workbook = load_workbook(self.workbook_path)
            worksheet = workbook.active
            if worksheet is None:
                logger.error("Workbook has no active sheet")
                return
                
            day_text = day.strftime("%Y-%m-%d")
            rows_to_keep = []
            for row in worksheet.iter_rows(values_only=True):
                if not row:
                    continue
                row_date = row[0]
                # Convert row_date to string if it's a datetime object
                if isinstance(row_date, datetime):
                    row_date_str = row_date.strftime("%Y-%m-%d")
                else:
                    row_date_str = str(row_date) if row_date else ""
                if row_date_str != day_text:
                    rows_to_keep.append(list(row))
            
            worksheet.delete_rows(1, worksheet.max_row)
            worksheet.append(["Date", "Time", "Student ID", "Full Name", "Status"])
            for row in rows_to_keep:
                worksheet.append(row)
            workbook.save(self.workbook_path)
            logger.debug(f"Cleared records for {day_text}")
        except OSError as e:
            logger.error(f"Error clearing records: {e}")

    def export_daily_summary(self, day: datetime, students: list[Student]) -> Path:
        """Export a daily attendance summary to a new workbook.
        
        Args:
            day: Date to create summary for
            students: List of all student objects for reference
            
        Returns:
            Path to the exported summary workbook
        """
        ensure_directories()
        output_path = DAILY_LOG_DIR / f"Daily_Attendance_{day.strftime('%Y-%m-%d')}.xlsx" # Revert: Keep in DAILY_LOG_DIR

        # Get records and create an efficient lookup
        daily_records = self.records_for_day(day)
        present_records_map = {record.student_id: record for record in daily_records}
        sorted_students = sorted(students, key=lambda s: s.full_name.lower())

        try:
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet()
            sheet.title = f"Attendance {day.strftime('%b %d, %Y')}"

            # --- Create Headers ---
            headers = ["#", "Name", "Check-In Time", "Present", "Absent"] # Revert: Keep this template
            sheet.append(headers)
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            # --- Fill Data ---
            # Revert: Keep the icon logic for daily
            for i, student in enumerate(sorted_students, start=1):
                record = present_records_map.get(student.student_id)
                if record:
                    check_in_time = record.recorded_at.strftime("%I:%M %p")
                    present_icon = "✔"
                    absent_icon = "✘"
                else:
                    check_in_time = ""
                    present_icon = "✘"
                    absent_icon = "✔"
                
                sheet.append([i, student.full_name, check_in_time, present_icon, absent_icon])
                # Center align the icons
                sheet.cell(row=i + 1, column=4).alignment = Alignment(horizontal="center")
                sheet.cell(row=i + 1, column=5).alignment = Alignment(horizontal="center")

            # --- Final Formatting ---
            _auto_size_columns(sheet) # Revert: Keep auto-sizing

            workbook.save(output_path)
            logger.info(f"Exported daily summary to {output_path}")
            return output_path
        except OSError as e:
            logger.error(f"Failed to export daily summary: {e}")
            raise

    def export_weekly_summary(self, day_in_week: datetime, students: list[Student]) -> Path:
        """Export a weekly attendance summary to a new workbook.
        
        Args:
            day_in_week: A day within the week to be exported.
            students: List of all student objects for reference.
            
        Returns:
            Path to the exported summary workbook.
        """
        ensure_directories()
        
        # Determine the start and end of the current week (Monday to Friday)
        today = day_in_week.date()
        start_of_week = today - timedelta(days=today.weekday()) # Monday
        end_of_week = start_of_week + timedelta(days=4) # Friday

        week_label = f"{start_of_week.strftime('%Y-%m-%d')}_{end_of_week.strftime('%Y-%m-%d')}"
        output_path = WEEKLY_LOG_DIR / f"Weekly_Attendance_{week_label}.xlsx" # Revert: Keep in WEEKLY_LOG_DIR

        presence_for_week = self._get_presence_for_range(datetime.combine(start_of_week, time.min), datetime.combine(end_of_week, time.max))
        sorted_students = sorted(students, key=lambda s: s.full_name.lower())

        try:
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet()
            sheet.title = f"Weekly Attendance {start_of_week.strftime('%b %d')}"

            # --- Create Headers ---
            week_days = [start_of_week + timedelta(days=i) for i in range(5)] # Monday to Friday
            headers = ["#", "Student Name", ""] + [d.strftime("%a %b %d") for d in week_days] # Revert: Keep this template
            sheet.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # --- Fill Data ---
            daily_present_counts = [0] * len(week_days)
            # Revert: Keep the icon logic for weekly
            for i, student in enumerate(sorted_students, start=1):
                row_data = [i, student.full_name, ""]
                
                for day_index, d in enumerate(week_days):
                    date_str = d.strftime("%Y-%m-%d")
                    is_present = date_str in presence_for_week and student.student_id in presence_for_week[date_str]
                    row_data.append("✔" if is_present else "✘")
                    if is_present:
                        daily_present_counts[day_index] += 1
                sheet.append(row_data)

            # --- Add Summary Row ---
            sheet.append([]) # Add a blank row for spacing
            total_students = len(sorted_students)
            # Since week_days is only Mon-Fri, we don't need the weekday check here.
            daily_present_display = [count if count > 0 else "" for count in daily_present_counts]
            daily_absent_display = [total_students - count if count > 0 else "" for count in daily_present_counts]
            
            present_summary_row = ["", "Total Present", ""] + daily_present_display
            absent_summary_row = ["", "Total Absent", ""] + daily_absent_display
            sheet.append(present_summary_row)
            sheet.append(absent_summary_row)

            # Style the summary rows
            for row in sheet.iter_rows(min_row=sheet.max_row - 1, max_row=sheet.max_row):
                for cell in row:
                    cell.font = header_font

            # --- Final Formatting --- # Revert: Keep auto-sizing and center alignment
            # Center align the daily status icons
            for row in sheet.iter_rows(min_row=2, min_col=4, max_col=len(week_days) + 3):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

            _auto_size_columns(sheet) # Auto-size all columns

            workbook.save(output_path)
            logger.info(f"Exported weekly summary to {output_path}")
            return output_path
        except OSError as e:
            logger.error(f"Failed to export weekly summary: {e}")
            raise

    def export_monthly_summary(self, day: datetime, students: list[Student]) -> Path:
        """Export a monthly attendance summary to a new workbook.
        
        Args:
            day: A day within the month to be exported
            students: List of all student objects for reference
            
        Returns:
            Path to the exported summary workbook
        """
        ensure_directories()
        year, month = day.year, day.month
        today = datetime.now().date()
        month_name = day.strftime("%B")
        output_path = MONTHLY_LOG_DIR / f"Monthly_Attendance_{year}_{month_name}.xlsx" # Revert: Keep in MONTHLY_LOG_DIR

        # Load all records for the month once for efficiency
        presence_by_day = self._get_monthly_presence(year, month)
        sorted_students = sorted(students, key=lambda s: s.full_name.lower())
        
        # Get all days in the month
        _, num_days = calendar.monthrange(year, month)
        days_in_month = [datetime(year, month, d) for d in range(1, num_days + 1)] # Revert: Keep datetime objects

        try:
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet()
            sheet.title = f"{month_name} {year} Attendance"

            # --- Create Headers ---
            headers = ["#", "Student Name", ""] + [d.strftime("%b %d") for d in days_in_month] # Revert: Keep this template
            sheet.append(headers)
            
            # Style headers
            header_font = Font(bold=True)
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # --- Fill Data ---
            weekend_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light red fill
            weekend_font = Font(color="9C0006") # Dark red font for header # Revert: Keep weekend font
            daily_present_counts = [0] * len(days_in_month)
            
            for i, student in enumerate(sorted_students, start=1):
                row_data = [i, student.full_name, ""]
                for day_index, d in enumerate(days_in_month):
                    date_str = d.strftime("%Y-%m-%d")
                    is_present = date_str in presence_by_day and student.student_id in presence_by_day[date_str]
                    day_is_in_past_or_present = d.date() <= today
                    day_has_started = bool(presence_by_day.get(date_str)) or d.date() < today
                    is_weekend = d.weekday() >= 5

                    mark = ""
                    if not is_weekend and day_is_in_past_or_present:
                        mark = "✔" if is_present else "✘" if day_has_started else ""
                    
                    row_data.append(mark)

                    if is_present:
                        daily_present_counts[day_index] += 1

                sheet.append(row_data)

            # --- Add Summary Row ---
            sheet.append([]) # Add a blank row for spacing
            total_students = len(sorted_students)
            present_summary = []
            absent_summary = []
            for i, d in enumerate(days_in_month):
                day_has_started = bool(presence_by_day.get(d.strftime("%Y-%m-%d"))) or d.date() < today
                show_summary = d.weekday() < 5 and d.date() <= today and day_has_started
                present_summary.append(str(daily_present_counts[i]) if show_summary else "")
                absent_summary.append(str(total_students - daily_present_counts[i]) if show_summary else "")
            present_summary_row = ["", "Total Present", ""] + present_summary
            absent_summary_row = ["", "Total Absent", ""] + absent_summary
            sheet.append(present_summary_row)
            sheet.append(absent_summary_row)

            # --- Final Formatting ---
            # Style summary and center-align icons # Revert: Keep styling
            summary_start_row = sheet.max_row - 1
            # Auto-size columns and apply weekend styling
            for i, col in enumerate(sheet.columns, start=1):  # i is 1-based column index
                if 3 < i <= (len(days_in_month) + 3):  # It's a day column (starts at col 4)
                    day_obj = days_in_month[i - 4]
                    if day_obj.weekday() >= 5:
                        for cell in col: # Apply weekend styling to the entire column
                            cell.fill = weekend_fill
                        # Set weekend header font color
                        sheet.cell(row=1, column=i).font = weekend_font
                    # Center align icons and summary counts
                    for cell_index, cell in enumerate(col, start=1):
                        if 1 < cell_index < summary_start_row: # Icon rows
                            cell.alignment = Alignment(horizontal="center") # Revert: Keep center alignment
                        elif cell_index >= summary_start_row: # Summary rows
                            # Apply bold font, but don't override weekend header font color
                            if not (day_obj.weekday() >= 5 and cell.row == 1):
                                cell.font = header_font
                    sheet.column_dimensions[get_column_letter(i)].width = 10 # Set day columns to a fixed small width # Revert: Keep fixed width
                elif i == 3: # The new blank column
                    sheet.column_dimensions[get_column_letter(i)].width = 3
                else: # Auto-size other columns
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    sheet.column_dimensions[get_column_letter(i)].width = max_length + 2

            workbook.save(output_path)
            logger.info(f"Exported monthly attendance to {output_path}")
            return output_path
        except OSError as e:
            logger.error(f"Failed to export summary: {e}")
            raise
